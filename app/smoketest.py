"""Go-live chat smoke-test runner (m35) — `python -m app.smoketest <client_id>`.

A /chat vegpontot hivja (a FUTO api-t), a tesztsort a pure
app.services.smoketest_battery adja, az eredmenybol ugyfelnek atadhato XLSX
riportot ir. Az utolso stdout-sor egy JSON osszegzo — az /smoketest endpoint
ebbol irja az event-metat (a /sync mintajara).

Kulon processzben fut (subprocess vagy `docker compose run`), ezert a szinkron
httpx nem blokkolja az api event loopjat.
"""
from __future__ import annotations

import argparse
import asyncio
import datetime as dt
import json
import sys

import httpx
from sqlalchemy import select

from app.core.db import SessionLocal
from app.core.settings import get_settings
from app.models.db_models import Tenant
from app.services.smoketest_battery import build_cases, evaluate, order_form_fields_of


# ---------------------------------------------------------------- tenant + KB
async def load_tenant(cid: str) -> dict:
    async with SessionLocal() as s:
        t = (await s.execute(select(Tenant).where(Tenant.client_id == cid))).scalar_one_or_none()
    if t is None:
        print(f"HIBA: nincs ilyen tenant: {cid}")
        sys.exit(2)
    return {
        "client_id": t.client_id,
        "platform": (t.platform or "").lower(),
        "plan": t.plan or "",
        "bot_name": getattr(t, "bot_name", "") or "",
        "welcome": getattr(t, "welcome_message", "") or "",
        "live_agent": bool(getattr(t, "live_agent_enabled", False)),
        "elallas": bool((getattr(t, "elallas_url", "") or "").strip()),
        "search_fb": bool(getattr(t, "search_fallback", False)),
    }


def kb_doc_count(cid: str) -> int:
    st = get_settings()
    url = st.qdrant_url.rstrip("/")
    key = getattr(st, "qdrant_api_key", "") or ""
    h = {"api-key": key} if key else {}
    coll = st.qdrant_collection
    cidf = {"key": "client_id", "match": {"value": cid}}
    try:
        tot = httpx.post(f"{url}/collections/{coll}/points/count",
                         json={"filter": {"must": [cidf]}, "exact": True},
                         headers=h, timeout=30).json()["result"]["count"]
        prod = httpx.post(f"{url}/collections/{coll}/points/count",
                          json={"filter": {"must": [cidf, {"key": "type", "match": {"value": "product"}}]},
                                "exact": True}, headers=h, timeout=30).json()["result"]["count"]
        return max(0, tot - prod)
    except Exception as e:  # noqa: BLE001
        print(f"figyelem: KB-count nem elerheto ({e})")
        return -1


# ---------------------------------------------------------------- /chat hivas
def run_case(base: str, cid: str, case: dict, idx: int, platform: str) -> dict:
    payload = {"client_id": cid, "message": case["kerdes"],
               "session_id": f"golive-{cid}-{idx}", "type": "message",
               "history": case.get("history") or []}
    reply, action, of_fields, err = "", None, None, ""
    try:
        r = httpx.post(f"{base}/chat", json=payload, timeout=60)
        r.raise_for_status()
        d = r.json()
        reply = (d.get("reply") or "").strip()
        action = d.get("action")
        of_fields = order_form_fields_of(d)
    except Exception as e:  # noqa: BLE001
        err = str(e)
    status, note = evaluate(case, action, of_fields, reply, err, platform)
    return {**case, "reply": reply or ("(ures valasz)" if not err else ""),
            "action": action, "status": status, "note": note}


# ---------------------------------------------------------------- XLSX riport
def write_xlsx(path: str, cfg: dict, rows: list[dict], base: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    ACCENT = "A08B6E"
    DARK = "3D3226"
    HEAD_FILL = PatternFill("solid", fgColor=ACCENT)
    META_FILL = PatternFill("solid", fgColor="F3EFE9")
    FB_FILL = PatternFill("solid", fgColor="FFF8E1")
    OK_FILL = PatternFill("solid", fgColor="E7F3E7")
    WARN_FILL = PatternFill("solid", fgColor="FCE9D6")
    ERR_FILL = PatternFill("solid", fgColor="F6D9D9")
    thin = Side(style="thin", color="D9CFC2")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    WRAP = Alignment(wrap_text=True, vertical="top")
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Chat go-live teszt"

    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = f"Chat go-live teszt — {cfg['client_id']}"
    c.font = Font(bold=True, size=16, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=DARK)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    ok = sum(1 for r in rows if r["status"] == "OK")
    warn = sum(1 for r in rows if r["status"] == "NEZD MEG")
    err = sum(1 for r in rows if r["status"] == "HIBA")
    manual = sum(1 for r in rows if r["status"] == "— kezi")
    funk = ["elo atadas: BE" if cfg["live_agent"] else "elo atadas: ki",
            "elallasi urlap: van" if cfg["elallas"] else "elallasi urlap: nincs",
            "bolti kereso: BE" if cfg["search_fb"] else "bolti kereso: ki",
            f"KB-hazirend doksi: {cfg['_kb']} chunk" if cfg["_kb"] >= 0 else "KB: n/a"]
    meta = [
        ("Platform", cfg["platform"]),
        ("Csomag", cfg["plan"]),
        ("Bot neve", cfg["bot_name"] or "(nincs)"),
        ("Udvozlo uzenet", cfg["welcome"] or "(nincs)"),
        ("Funkciok", " · ".join(funk)),
        ("Teszt idopontja", dt.datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("API", base),
        ("Eredmeny", f"osszesen {len(rows)} teszt — OK: {ok} · nezd meg: {warn} · hiba: {err} · kezi ertekeles: {manual}"),
    ]
    row = 3
    for k, v in meta:
        ws.cell(row=row, column=1, value=k).font = Font(bold=True, color=DARK)
        ws.cell(row=row, column=1).fill = META_FILL
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
        cc = ws.cell(row=row, column=2, value=str(v))
        cc.alignment = Alignment(wrap_text=True, vertical="center")
        cc.fill = META_FILL
        row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    leg = ws.cell(row=row, column=1,
                  value=("Jelmagyarazat — OK: automatikusan ellenorzott, megfelelt · "
                         "NEZD MEG: automatikus ellenorzes elterest jelzett, nezd at a valaszt · "
                         "— kezi: tartalmi ertekelest igenyel (te/az ugyfel dontse el, jo-e). "
                         "Toltsd ki a ket sarga oszlopot (Megfelelo? I/N + Megjegyzes), es kuldd vissza a finomitashoz."))
    leg.font = Font(italic=True, color=DARK, size=10)
    leg.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[row].height = 42
    row += 2

    headers = ["#", "Kategoria", "Teszt celja", "Elvart viselkedes", "Kerdes (amit a bot kapott)",
               "Bot valasza", "Auto-ellenorzes", "Megfelelo? (I/N)", "Megjegyzes / javitando valasz"]
    hr = row
    for j, htxt in enumerate(headers, start=1):
        cell = ws.cell(row=hr, column=j, value=htxt)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEAD_FILL
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[hr].height = 30
    row += 1

    for i, r in enumerate(rows, start=1):
        vals = [i, r["kat"], r["cel"], r["elvart"], r["kerdes"], r["reply"], r["status"], "", ""]
        for j, v in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=j, value=v)
            cell.border = BORDER
            cell.alignment = CENTER if j in (1, 7, 8) else WRAP
            if j == 7:
                if r["status"] == "OK":
                    cell.fill = OK_FILL
                elif r["status"] == "NEZD MEG":
                    cell.fill = WARN_FILL
                elif r["status"] == "HIBA":
                    cell.fill = ERR_FILL
                if r["note"]:
                    cell.value = f"{r['status']}\n{r['note']}"
            if j in (8, 9):
                cell.fill = FB_FILL
        row += 1

    widths = {1: 4, 2: 15, 3: 22, 4: 34, 5: 30, 6: 60, 7: 18, 8: 12, 9: 34}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = ws.cell(row=hr + 1, column=1)
    ws.auto_filter.ref = f"A{hr}:I{row - 1}"
    wb.save(path)


# ---------------------------------------------------------------- main
def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("client_id")
    ap.add_argument("--base", default="http://localhost:8000")
    ap.add_argument("--out", default="/reports")
    ap.add_argument("--outfile", default="", help="teljes celutvonal; ha ures, --out + idopecset")
    args = ap.parse_args()
    cid = args.client_id.strip().lower()

    cfg = asyncio.run(load_tenant(cid))
    cfg["_kb"] = kb_doc_count(cid)
    print(f"[{cid}] platform={cfg['platform']} plan={cfg['plan']} live={cfg['live_agent']} "
          f"elallas={cfg['elallas']} kereso={cfg['search_fb']} KB={cfg['_kb']}")

    cases = build_cases(cfg)
    print(f"[{cid}] {len(cases)} teszteset futtatasa a {args.base} ellen...")

    rows = []
    for i, case in enumerate(cases, start=1):
        r = run_case(args.base, cid, case, i, cfg["platform"])
        rows.append(r)
        print(f"  {i:2d}. [{r['status']:9s}] {r['kat']:16s} | {r['cel']}", flush=True)

    outfile = args.outfile or f"{args.out}/chat-teszt-{cid}-{dt.datetime.now().strftime('%Y%m%d-%H%M')}.xlsx"
    write_xlsx(outfile, cfg, rows, args.base)
    print(f"XLSX kesz: {outfile}")

    ok = sum(1 for r in rows if r["status"] == "OK")
    warn = sum(1 for r in rows if r["status"] == "NEZD MEG")
    err = sum(1 for r in rows if r["status"] == "HIBA")
    manual = sum(1 for r in rows if r["status"] == "— kezi")
    # az UTOLSO sor JSON — az /smoketest endpoint event-metaja ebbol keszul
    print(json.dumps({"report": outfile.rsplit("/", 1)[-1], "total": len(rows),
                      "ok": ok, "warn": warn, "err": err, "manual": manual}))


if __name__ == "__main__":
    main()
