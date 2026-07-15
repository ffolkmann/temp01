"""Teljes beszelgetes-export XLSX (m48) - a stat.html "Letoltes (XLSX)" gombja.

Ket munkalap: session-onkent osszevont beszelgetesek (legfrissebb felul) +
nyers turn-naplo. Forras: a `messages` naplo (30 nap retention). Szandekosan
minimal fuggoseg (openpyxl + stdlib), az app-ot nem importalja - a teszt
fajlbol tolti (tests/test_conversations_export_m48.py).
"""

from __future__ import annotations

import io
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

BUDAPEST = ZoneInfo("Europe/Budapest")

_CELL_LIMIT = 20000  # Excel cella-limit 32767 - biztonsagi vagas alatta

_SHEET_SESS = "Besz\u00e9lget\u00e9sek"
_SHEET_RAW = "Nyers napl\u00f3"

_HDR_SESS = ["Session ID", "Kezdet", "Utols\u00f3 \u00fczenet", "\u00dczenetek",
             "Besz\u00e9lget\u00e9s"]
_HDR_RAW = ["Id\u0151pont", "Session ID", "K\u00e9rd\u00e9s", "V\u00e1lasz", "M\u0171velet"]


def _fmt_ts(dt) -> str:
    if dt is None:
        return ""
    try:
        return dt.astimezone(BUDAPEST).strftime("%Y-%m-%d %H:%M")
    except Exception:
        return str(dt)


def _fmt_transcript(turns) -> str:
    """turnok -> olvashato atirat: [ts] U: kerdes / B: valasz blokkok."""
    parts = []
    for t in (turns or []):
        q = (t.get("question") or "").strip()
        a = (t.get("answer") or "").strip()
        ts = _fmt_ts(t.get("created_at"))
        if q:
            parts.append("[%s] U: %s" % (ts, q))
        if a:
            parts.append("B: %s" % a)
        parts.append("")
    out = "\n".join(parts).strip()
    if len(out) > _CELL_LIMIT:
        out = out[:_CELL_LIMIT] + "\n... [v\u00e1gva]"
    return out


def build_conversations_xlsx(rows: list[dict]) -> bytes:
    """rows: messages sorok (session_id, question, answer, action, created_at)
    session_id, created_at, id szerint rendezve (a raw lap ebben a sorrendben ir)."""
    sessions: dict[str, dict] = {}
    for r in rows:
        sid = r.get("session_id") or ""
        g = sessions.get(sid)
        if g is None:
            g = sessions[sid] = {"first": r.get("created_at"), "last": r.get("created_at"),
                                 "n": 0, "turns": []}
        g["n"] += 1
        g["last"] = r.get("created_at")
        g["turns"].append(r)

    # legfrissebb session felul (a formazott ts lexikografikusan idorendes)
    ordered = sorted(sessions.items(), key=lambda kv: _fmt_ts(kv[1]["last"]), reverse=True)

    wb = Workbook()
    ws = wb.active
    ws.title = _SHEET_SESS
    ws.append(_HDR_SESS)
    for sid, g in ordered:
        ws.append([sid, _fmt_ts(g["first"]), _fmt_ts(g["last"]), g["n"],
                   _fmt_transcript(g["turns"])])

    ws2 = wb.create_sheet(_SHEET_RAW)
    ws2.append(_HDR_RAW)
    for r in rows:
        ws2.append([_fmt_ts(r.get("created_at")), r.get("session_id") or "",
                    r.get("question") or "", r.get("answer") or "", r.get("action") or ""])

    wrap = Alignment(wrap_text=True, vertical="top")
    for sheet, widths, wraps in ((ws, [34, 17, 17, 10, 100], (5,)),
                                 (ws2, [17, 34, 55, 70, 16], (3, 4))):
        bold = Font(bold=True)
        for c in sheet[1]:
            c.font = bold
        for i, w in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(i)].width = w
        for rcells in sheet.iter_rows(min_row=2):
            for ci in wraps:
                rcells[ci - 1].alignment = wrap
        sheet.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
