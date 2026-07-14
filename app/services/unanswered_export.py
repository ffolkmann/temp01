"""Megvalaszolatlan kerdesek XLSX exportja (m44+m45) - a stat.html Letoltes gombja.

Ket munkalap: csoportositott kerdesek (elofordulas szerint) + nyers naplo,
mindketto beszelgetes-kontextussal (messages naplo, 30 nap retention).
Szandekosan minimal fuggoseg (openpyxl + stdlib), az app-ot nem importalja -
a teszt fajlbol tolti (lasd tests/test_unanswered_export_m44.py).
"""

from __future__ import annotations

import io
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

BUDAPEST = ZoneInfo("Europe/Budapest")

_CELL_LIMIT = 20000  # Excel cella-limit 32767 - biztonsagi vagas alatta

REASON_HU = {
    "low_score": "nincs tal\u00e1lat",
    "collect_lead": "lead-k\u00e9r\u00e9s",
    "order_form": "rendel\u00e9s-\u0171rlap",
}

_SHEET_GROUPED = "K\u00e9rd\u00e9sek (csoportos\u00edtva)"
_SHEET_RAW = "Nyers napl\u00f3"

_HDR_GROUPED = ["K\u00e9rd\u00e9s", "El\u0151fordul\u00e1s", "Score", "Okok",
                "Utols\u00f3 el\u0151fordul\u00e1s", "Session ID-k (max 5)",
                "Legut\u00f3bbi besz\u00e9lget\u00e9s (kontextus)"]
_HDR_RAW = ["Id\u0151pont", "K\u00e9rd\u00e9s", "Score", "Okok", "Session ID",
            "Besz\u00e9lget\u00e9s (kontextus)"]

_NO_LOG = "nincs napl\u00f3 (30+ nap)"


def _fmt_ts(dt) -> str:
    if dt is None:
        return ""
    try:
        return dt.astimezone(BUDAPEST).strftime("%Y-%m-%d %H:%M")
    except Exception:
        return str(dt)


def _hu_reasons(reasons) -> str:
    return ", ".join(REASON_HU.get(r, r) for r in (reasons or []))


def _fmt_transcript(turns) -> str:
    """messages-turnok -> olvashato atirat: [ts] U: kerdes / B: valasz blokkok."""
    parts = []
    for t in (turns or []):
        q = (t.get("question") or "").strip()
        a = (t.get("answer") or "").strip()
        ts = _fmt_ts(t.get("created_at"))
        if q:
            parts.append("[%s] U: %s" % (ts, q))
        if a:
            parts.append("B: %s" % a)
        parts.append("")
    out = "\n".join(parts).strip()
    if len(out) > _CELL_LIMIT:
        out = out[:_CELL_LIMIT] + "\n... [v\u00e1gva]"
    return out


def _ctx(sid, transcripts) -> str:
    """Kontextus-cella: atirat, ha van naplo; jelzes, ha lejart; ures, ha nincs sid."""
    if not sid:
        return ""
    turns = (transcripts or {}).get(sid)
    if not turns:
        return _NO_LOG
    return _fmt_transcript(turns)


def build_unanswered_xlsx(rows: list[dict], transcripts: dict | None = None) -> bytes:
    """rows: unanswered sorok created_at DESC sorrendben
    (question, score, reasons, session_id, created_at kulcsokkal).
    transcripts: session_id -> messages-turnok (question, answer, created_at) idorendben."""
    groups: dict[str, dict] = {}
    for r in rows:
        q = r.get("question") or ""
        g = groups.get(q)
        if g is None:  # elso elofordulas = legutobbi (DESC) -> score + last_ts innen
            g = groups[q] = {"count": 0, "last_ts": r.get("created_at"),
                             "score": r.get("score"), "reasons": set(), "sessions": []}
        g["count"] += 1
        sid = r.get("session_id")
        if sid and sid not in g["sessions"] and len(g["sessions"]) < 5:
            g["sessions"].append(sid)
        for rs in (r.get("reasons") or []):
            g["reasons"].add(rs)

    grouped = sorted(groups.items(),
                     key=lambda kv: (kv[1]["count"], _fmt_ts(kv[1]["last_ts"])),
                     reverse=True)

    wb = Workbook()
    ws = wb.active
    ws.title = _SHEET_GROUPED
    ws.append(_HDR_GROUPED)
    for q, g in grouped:
        ws.append([
            q, g["count"],
            round(float(g["score"]), 4) if g["score"] is not None else None,
            _hu_reasons(sorted(g["reasons"])),
            _fmt_ts(g["last_ts"]),
            ", ".join(g["sessions"]),
            _ctx(g["sessions"][0] if g["sessions"] else "", transcripts),
        ])

    ws2 = wb.create_sheet(_SHEET_RAW)
    ws2.append(_HDR_RAW)
    for r in rows:
        ws2.append([
            _fmt_ts(r.get("created_at")),
            r.get("question") or "",
            round(float(r["score"]), 4) if r.get("score") is not None else None,
            _hu_reasons(r.get("reasons")),
            r.get("session_id") or "",
            _ctx(r.get("session_id"), transcripts),
        ])

    wrap = Alignment(wrap_text=True, vertical="top")
    for sheet, widths, qcol in ((ws, [60, 12, 9, 26, 18, 46, 90], 1),
                                (ws2, [18, 60, 9, 26, 40, 90], 2)):
        bold = Font(bold=True)
        for c in sheet[1]:
            c.font = bold
        for i, w in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(i)].width = w
        last = len(widths)
        for rcells in sheet.iter_rows(min_row=2):
            rcells[qcol - 1].alignment = wrap
            rcells[last - 1].alignment = wrap
        sheet.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
