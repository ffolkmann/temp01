"""m44+m45 - build_unanswered_xlsx: fajlbol toltve (suite-konvencio).

Vedekezes: ha korabbi teszt fake openpyxl-t ultetett a sys.modules-ba,
save/restore-ral felretesszuk a betoltes idejere.
"""
import importlib.util
import io
import sys
from datetime import datetime, timezone
from pathlib import Path

MOD_PATH = Path(__file__).resolve().parents[1] / "app" / "services" / "unanswered_export.py"

_SAVED = {}


def setup_module():
    for name in list(sys.modules):
        if name == "openpyxl" or name.startswith("openpyxl."):
            _SAVED[name] = sys.modules.pop(name)


def teardown_module():
    for name in list(sys.modules):
        if name == "openpyxl" or name.startswith("openpyxl."):
            sys.modules.pop(name)
    sys.modules.update(_SAVED)


def _load():
    spec = importlib.util.spec_from_file_location("unanswered_export_m44", str(MOD_PATH))
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def _dt(s):
    return datetime.fromisoformat(s).replace(tzinfo=timezone.utc)


def test_grouped_and_raw():
    mod = _load()
    rows = [
        {"question": "van garancia?", "score": 0.31, "reasons": ["low_score"],
         "session_id": "s3", "created_at": _dt("2026-07-14T10:00:00")},
        {"question": "mikor nyittok?", "score": None, "reasons": ["collect_lead"],
         "session_id": "s2", "created_at": _dt("2026-07-13T09:00:00")},
        {"question": "van garancia?", "score": 0.28, "reasons": ["low_score", "order_form"],
         "session_id": "s1", "created_at": _dt("2026-07-12T08:00:00")},
    ]
    data = mod.build_unanswered_xlsx(rows)
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["K\u00e9rd\u00e9sek (csoportos\u00edtva)", "Nyers napl\u00f3"]
    g = wb[wb.sheetnames[0]]
    vals = list(g.iter_rows(min_row=2, values_only=True))
    assert len(vals) == 2
    # count DESC: "van garancia?" (2x) elol; score a LEGFRISSEBB sorbol (0.31)
    assert vals[0][0] == "van garancia?" and vals[0][1] == 2 and vals[0][2] == 0.31
    assert "nincs tal\u00e1lat" in vals[0][3] and "rendel\u00e9s-\u0171rlap" in vals[0][3]
    assert vals[0][4] == "2026-07-14 12:00"  # UTC 10:00 -> Bp CEST 12:00
    assert vals[0][5] == "s3, s1"
    assert vals[1][0] == "mikor nyittok?" and vals[1][2] is None
    r = wb[wb.sheetnames[1]]
    raws = list(r.iter_rows(min_row=2, values_only=True))
    assert len(raws) == 3
    assert raws[0][0] == "2026-07-14 12:00"
    assert raws[0][1] == "van garancia?" and raws[0][4] == "s3"


def test_empty():
    mod = _load()
    data = mod.build_unanswered_xlsx([])
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data))
    g = wb[wb.sheetnames[0]]
    assert list(g.iter_rows(min_row=2, values_only=True)) == []


def test_transcript_context():
    mod = _load()
    rows = [
        {"question": "van garancia?", "score": 0.31, "reasons": ["low_score"],
         "session_id": "s3", "created_at": _dt("2026-07-14T10:00:00")},
        {"question": "hol a bolt?", "score": 0.2, "reasons": ["low_score"],
         "session_id": None, "created_at": _dt("2026-07-13T09:00:00")},
        {"question": "van garancia?", "score": 0.28, "reasons": ["low_score"],
         "session_id": "s1", "created_at": _dt("2026-07-12T08:00:00")},
    ]
    transcripts = {
        "s3": [
            {"question": "milyen nyomtatok vannak?", "answer": "Tobbfele is.",
             "created_at": _dt("2026-07-14T09:58:00")},
            {"question": "van garancia?", "answer": "",
             "created_at": _dt("2026-07-14T10:00:00")},
        ],
    }
    data = mod.build_unanswered_xlsx(rows, transcripts)
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data))
    g = wb[wb.sheetnames[0]]
    vals = list(g.iter_rows(min_row=2, values_only=True))
    ctx = vals[0][6]  # "van garancia?" csoport -> s3 atirata
    assert "U: milyen nyomtatok vannak?" in ctx
    assert "B: Tobbfele is." in ctx
    assert "[2026-07-14 11:58]" in ctx  # UTC 09:58 -> Bp 11:58
    r = wb[wb.sheetnames[1]]
    raws = list(r.iter_rows(min_row=2, values_only=True))
    assert "U: milyen nyomtatok vannak?" in raws[0][5]          # s3
    assert raws[1][5] in (None, "")                              # session nelkul
    assert raws[2][5] == "nincs napl\u00f3 (30+ nap)"            # s1: nincs naplo


def test_transcript_truncate():
    mod = _load()
    rows = [{"question": "q", "score": 0.1, "reasons": ["low_score"],
             "session_id": "sX", "created_at": _dt("2026-07-14T10:00:00")}]
    transcripts = {"sX": [{"question": "q", "answer": "x" * 25000,
                           "created_at": _dt("2026-07-14T10:00:00")}]}
    data = mod.build_unanswered_xlsx(rows, transcripts)
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data))
    r = wb[wb.sheetnames[1]]
    raws = list(r.iter_rows(min_row=2, values_only=True))
    cell = raws[0][5]
    assert cell.endswith("[v\u00e1gva]")
    assert len(cell) <= 20000 + 20
