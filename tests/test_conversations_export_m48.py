"""m48 - beszelgetes-export XLSX (build_conversations_xlsx). Fajlbol toltve,
mint test_unanswered_export_m44 (nincs app-import, csak openpyxl + stdlib)."""

import importlib.util
import io
import pathlib
from datetime import datetime, timezone

from openpyxl import load_workbook

ROOT = pathlib.Path(__file__).resolve().parents[1]
_p = ROOT / "app" / "services" / "conversations_export.py"
_spec = importlib.util.spec_from_file_location("conversations_export_m48_under_test", _p)
_ce = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(_ce)


def _ts(h, m=0):
    return datetime(2026, 7, 15, h, m, tzinfo=timezone.utc)


def _rows():
    return [
        {"session_id": "s1", "question": "kerdes1", "answer": "valasz1",
         "action": None, "created_at": _ts(8)},
        {"session_id": "s1", "question": "kerdes2", "answer": "valasz2",
         "action": "collect_lead", "created_at": _ts(8, 5)},
        {"session_id": "s2", "question": "masik", "answer": "valasz",
         "action": None, "created_at": _ts(9)},
    ]


def _load(data):
    return load_workbook(io.BytesIO(data))


def test_sessions_grouped_and_ordered():
    wb = _load(_ce.build_conversations_xlsx(_rows()))
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 2
    assert rows[0][0] == "s2"  # legfrissebb session felul
    assert rows[1][0] == "s1"
    assert rows[1][3] == 2     # uzenetek szama
    assert "U: kerdes1" in rows[1][4] and "B: valasz2" in rows[1][4]


def test_raw_sheet_has_all_turns():
    wb = _load(_ce.build_conversations_xlsx(_rows()))
    ws = wb.worksheets[1]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 3
    assert rows[0][1] == "s1" and rows[0][2] == "kerdes1"
    assert rows[1][4] == "collect_lead"


def test_empty_input_headers_only():
    wb = _load(_ce.build_conversations_xlsx([]))
    assert list(wb.worksheets[0].iter_rows(min_row=2, values_only=True)) == []
    assert list(wb.worksheets[1].iter_rows(min_row=2, values_only=True)) == []


def test_transcript_cell_capped():
    rows = [{"session_id": "s1", "question": "x" * 3000, "answer": "y" * 3000,
             "action": None, "created_at": _ts(8, i)} for i in range(10)]
    wb = _load(_ce.build_conversations_xlsx(rows))
    cell = list(wb.worksheets[0].iter_rows(min_row=2, values_only=True))[0][4]
    assert len(cell) <= 20000 + 40
    assert "[v\u00e1gva]" in cell
